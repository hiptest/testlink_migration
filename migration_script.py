# -- utf-8 --

from lxml import etree

from bs4 import BeautifulSoup

from openpyxl import Workbook

import sys

import getopt

import logging

 

 

HEADERS = ['Test ID', 'Test name', 'Test description', 'Test tags', 'Pre-conditions', 'Steps', 'Result']

NOT_ALLOWED_CHAR_IN_EXCEL_WS_NAME = ["\"", "/", "*", "[", "]", ":", "?"]

 

def clean_text(text:str) -> str:

    local_text = BeautifulSoup(text, features="lxml")

    return local_text.get_text().replace('\t','').lstrip('\n').rstrip('\n')

 

def create_header(ws):

    for col, text in enumerate(HEADERS):

        ws.cell(row=1, column=col+1).value = text

 

def resize_columns(wb):

    for ws in wb.worksheets:

        logging.debug("resize columns for sheet '{}'".format(ws.title))

        ws.column_dimensions[ws.cell(row=1, column=2).column].width = 50

        ws.column_dimensions[ws.cell(row=1, column=5).column].width = 50

        ws.column_dimensions[ws.cell(row=1, column=6).column].width = 50

        ws.column_dimensions[ws.cell(row=1, column=7).column].width = 50

 

def read_summary(line:list, elem):

    line[2] = clean_text(elem.text)

 

#

# Construit la ligne contenant les colonnes steps et result

# Exception : la premiÃ¨re ligne doit contenir les colonnes summary et preconditions

#

 

def read_testcase(row:int, ws, test_case):

    line =  ['', '', '', '', '','','']

    line[1] = test_case.attrib['name']

    logging.debug("test_case = '{}'".format(line[1]))

    for elem in test_case:

        if elem.tag == "summary":

            read_summary(line, elem)

 

        elif elem.tag == "preconditions":

            read_preconditions(line, elem)

 

        elif elem.tag == "steps":

            for index, step in enumerate(elem.getchildren()):

                row += 1

                logging.debug("index : {} ; row : '{}'".format(index, row))

 

                if index == 0:

                    read_step(line, step)

                else:

                    line = [ '' ] * 7

                    read_step(line, step)

 

                for col, text in enumerate(line):

                    ws.cell(row=row,column=col+1).value = text

 

    return row

 

def read_step(line:list, step):

    for elem in step:

        if elem.tag == "actions":

            line[5] = clean_text(elem.text)

        elif elem.tag == "expectedresults":

            line[6] = clean_text(elem.text)

 

def read_preconditions(line:list, preconditions):

    line[4] = clean_text(preconditions.text)

 

def read_testsuite(row:int, ws, path:list, test_suite, isFirstLevel:bool=False):

    if not isFirstLevel:

        path.append(test_suite.attrib['name'])

        logging.debug("row = {} path : {}".format(row, path))

    ws.cell(row=row, column=1).value = "/".join(path)

    elems = test_suite.getchildren()

    row += 2            # Ligne blanche aprÃ¨s l'affichage du path

    for elem in elems:

        if elem.tag == 'testcase':

            row = read_testcase(row, ws, elem)

            row += 1

    for elem in elems:

        if elem.tag == 'testsuite':

            logging.debug("test_suite = {}".format(elem.attrib['name']))

            row = read_testsuite(row , ws, path, elem, False)

            path.pop()

 

    return row

 

 

def clean_ws_title(text:str):

    logging.debug("clean_ws_title")

 

    # Truncate name if too long

    truncated_test_suite_name = text[0:MAX_SHEET_NAME_LENGTH]

    if len(truncated_test_suite_name) < len(text):

        logging.warning(

                "Due to excel limitations, test_suite '{}' has been truncated to '{}'".format(

                        test_suite_name,

                        truncated_test_suite_name,

                ),

        )

 

    # Remove non authorized characters

    clean_text = truncated_test_suite_name

    for char in NOT_ALLOWED_CHAR_IN_EXCEL_WS_NAME:

        clean_text = clean_text.replace(char, '')

    if (len(clean_text) != len(truncated_test_suite_name)):

        logging.warning(

                "Due to excel limitations, test_suite '{}' has been renamed to '{}'".format(

                        truncated_test_suite_name,

                        clean_text,

                ),

        )

 

    # Return cleaned title

    return clean_text

 

 

def usage():

    # On affiche l'aide

    logging.info("python {} -i <inputfilepathname>".format(sys.argv[0]))

 

 

def disp_usage_and_exit():

    usage()

    sys.exit(2)

 

if __name__ == '__main__':

 

    logging.basicConfig(level=logging.INFO)

    MAX_SHEET_NAME_LENGTH = 31

 

    try:

        opts, args = getopt.getopt(

                sys.argv[1:],

                "hi:v",

                [

                    "help",

                    "inputfilepathname=",

                    "verbose",

                ],

        )

    except getopt.GetoptError as err:

        # Affiche l'aide et quitte le programme

        logging.error(err)  # va afficher l'erreur en anglais

        disp_usage_and_exit()

 

    if not len(sys.argv) > 1:

        disp_usage_and_exit()

 

    input_file_path_name = ""

    output_file_path_name = "./export.xlsx"

 

    for option, arg in opts:

        if option in ("-v", "--verbose"):

            logging.getLogger().setLevel(logging.DEBUG)

        elif option in ("-h", "--help"):

            usage()

            sys.exit()

        elif option in ("-i", "--inputfilepathname"):

            input_file_path_name = arg

 

    if input_file_path_name == "":

        disp_usage_and_exit()

 

    #Â --- ouverture du fichier

    with open(input_file_path_name, encoding='utf-8') as obj:

        xml = obj.read().encode('utf-8')

 

    #Â --- chargement du fichier xml

    root = etree.fromstring(xml)

 

    # --- creation de la structure spreadsheet

    wb = Workbook()

    first_suite_element_flag = True

 

    #Â ---- execution (recursive) ----

    for elem in root.getchildren():

        if elem.tag == "testsuite":

            logging.debug("test_suite = {}".format(elem.attrib['name']))

            test_suite_name = elem.attrib['name']

            cleaned_test_suite_name = clean_ws_title(test_suite_name)

 

            if first_suite_element_flag:

                first_suite_element_flag = False

                ws = wb.active

                ws.title = cleaned_test_suite_name

                create_header(wb[cleaned_test_suite_name])

            else:

                ws = wb.create_sheet(cleaned_test_suite_name)

                create_header(ws)

            row = 2

            row = read_testsuite(row, ws, [], elem, True) + 3

 

    # --- Resize columns

    resize_columns(wb)

 

    # --- save spreadsheet

    wb.save(output_file_path_name)
